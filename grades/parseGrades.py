"""Parse grades manually entered in yaml format.
Example: 

---
grades:
- matrikel: 1234567
  punkte: 
  - blatt: 1
    punkte: 5
    tutor: jb
  - blatt: 2
    punkte: 17
    tutor: dp
- matrikel: 2345678
  punkte: 
  - blatt: 1
    punkte: 4
    tutor: as
  - blatt: 1
    punkte: 7
    tutor: hk
- matrikel: 2345678
  punkte: 
  - blatt: 1
    punkte: 5
    tutor: as
  - blatt: 7
    punkte: 9
    tutor: hk
...

This script tries to cleanup and collect duplicates, but obvisouly it makes 
very little sense to have duplicates inserted. So watch out what you are doing! 

These points are MAXED with anything that comes out of the studentresult app.  

studentresult yaml output produces a  dict of the following structure: 
a dict over matrikelnummer, value a dict over hue0x, value a dict of group, mail, point 

'6501154':
  hue01:
    group: gp1_16_243
    mail: alexander.isaak@gmx.de
    point: 18.0
  hue02:
    group: gp1_16_243
    mail: alis@mail.uni-paderborn.de
    point: 20.0
  hue03:
    group: gp1_16_243
    mail: alis@mail.uni-paderborn.de
    point: 18.0
  hue04:
    group: gp1_16_243
    mail: alis@mail.uni-paderborn.de
    point: 17.0
  hue05:
    group: gp1_16_243
    mail: alis@mail.uni-paderborn.de
    point: 16.0
  hue06:
    group: null
    mail: ''
    point: 0
  hue07:
    group: null
    mail: ''
    point: 0


"""


import yaml
import sys
from collections import defaultdict
from pprint import pprint as pp
from jinja2 import Template
from openpyxl import Workbook
import matplotlib.pyplot as plt
import argparse

############


# how many assignments are there? 
max_num_assignments = 13

############

def getManual(gradefile):
    with open(gradefile) as y:
        grades = yaml.load(y)

    grades = grades['grades']
    # print(grades)

    # cleaup possible duplicates

    finalgrades = {}
    for m in set([x['matrikel'] for x in grades]):
        mgrades = [x
                    for x in grades
                    if x['matrikel'] == m]

        mfinal = defaultdict(int)

        for p in mgrades:
            for b in p['punkte']:
                mfinal[b['blatt']] += b['punkte']

        finalgrades[m] = mfinal

    return finalgrades

###########

def getFormgrades(gradefile):
    with open(gradefile) as y:
        grades = yaml.load(y)

    # sanitize the assignment names to match the manual format
    r = {}
    
    for matrikel, assignment in grades.items():
        # for aa in assignments:
        #     # print(aa)
        #     aa['assignment'] = int(aa['assignment'][3:])
        #     aa['manual'] = 0

        try:
            m = int(matrikel)
            assert(m> 1000000)
        except Exception as e:
            # print("formgrade: matrikel not integer: ", e, matrikel)
            pass 
        else:
            r[m] = {}
            for a, v in assignment.items():
                v['manual'] = 0
                r[m][int(a[3:])] = v

    # pp(r)
    return r

############
# probably no longer needed: 
def cleanup(formgrades):
    fg = filter(lambda x: x['matrikelnr'].isnumeric(),
                    formgrades
                    )
    return list(fg)

############

def merge(formgrades, manualgrades):

    # pp(formgrades)

    # Add manually graded points into formgrade structure 
    for matrikel, blaetter in manualgrades.items():
        # print("=====")
        # print(matrikel, blaetter)
        if matrikel not in formgrades:
            formgrades[matrikel] = {}
        fgmatrikel = formgrades[matrikel]
        # pp(fgmatrikel)
        
        for blatt, punkte in blaetter.items():
            if blatt not in fgmatrikel:
                fgmatrikel[blatt] = {'group': "manually added",
                                     'mail': 'N/A',
                                         'manual': 0,
                                         'point': 0}
                
            fgmatrikel[blatt]['manual'] = punkte

        # pp(fgmatrikel)


    # add manual and automatic points together: 
    for matrikel, assignments  in formgrades.items():
        for aa, values in assignments.items():
            if values['manual']>0 and  values['point']>0:
                print("both manual and automatic: ", matrikel, aa, values)
                
            values['total'] = max(values['manual'], values['point'])

    
    # pp(formgrades)
                
############

def texescape(s):
    if s:
        
        tex_replacements = [
            ('{', r'\{'),
            ('}', r'\}'),
            ('[', r'{[}'),
            (']', r'{]}'),
            ('\\', r'\textbackslash{}'),
            ('$', r'\$'),
            ('%', r'\%'),
            ('&', r'\&'),
            ('#', r'\#'),
            ('_', r'\_'),
            ('~', r'\textasciitilde{}'),
            ('^', r'\textasciicircum{}'),
            ]

        mapping = dict((ord(char), rep) for char, rep in tex_replacements)

        s = s.translate(mapping)
    return s
        
def escape(formgrades, fields):
    """Latex-escape text in fields of assignments"""

    for matrikel, assignments in formgrades.items():
        # student['matrikelnr'] = texescape(student['matrikelnr'])
        # assignments = student['assignments']
        for aa, values in assignments.items():
            for f in fields:
                values[f] = texescape(values[f])
            

############

def output(formgrades, succeeded=None):

    template = Template(r"""
\documentclass{article}
\usepackage[utf8]{inputenc}
\usepackage{graphicx}
\usepackage{hyperref}
\usepackage[ngerman]{babel}
\begin{document}
\title{Übersicht erzielte Punkte, GP1, WS 16/17, UPB}
\maketitle
\tableofcontents
\section{Punkteübersicht}
\begin{description}
{% for matrikelnr, assignments in grades.items() %} 
\item[Matrikelnummer {{ matrikelnr }} ]~\\
% Anzahl Blätter über Schwellwert für Klausurzulassung: 
\begin{description}
{% for a, v in assignments.items() %}
\item[Hausblatt {{ a }}: ]
\begin{description}
\item[Gruppe:] {{ v.group }}
\item[Punkte regulär:] {{ v.point }}
\item[Punkte manuell:] {{ v.manual }}
\item[Punkte gesamt:] {{ v.total }}
\end{description}
{% endfor %}
\end{description}
{% endfor %}
\end{description}

\section{Klausurzulassung}

Achtung, ohne Gewähr! 

\begin{itemize}
{% for matrikel, punkte in succeeded %} 
\item {{ matrikel }} , Gesamtpunkte: {{ punkte }}
{% endfor %}
\end{itemize}


\section{Statistiken}

Statistiken über Matrikelnummern, nicht über Gruppen! 

\subsection{Hausblatt 1}
\includegraphics[width=0.8\textwidth]{h1.pdf}
\subsection{Hausblatt 2}
\includegraphics[width=0.8\textwidth]{h2.pdf}
\subsection{Hausblatt 3}
\includegraphics[width=0.8\textwidth]{h3.pdf}
\subsection{Hausblatt 4}
\includegraphics[width=0.8\textwidth]{h4.pdf}
\subsection{Hausblatt 5}
\includegraphics[width=0.8\textwidth]{h5.pdf}
\subsection{Hausblatt 6}
\includegraphics[width=0.8\textwidth]{h6.pdf}
\subsection{Hausblatt 7}
\includegraphics[width=0.8\textwidth]{h7.pdf}
\subsection{Hausblatt 8}
\includegraphics[width=0.8\textwidth]{h8.pdf}
\subsection{Hausblatt 9}
\includegraphics[width=0.8\textwidth]{h9.pdf}
\subsection{Hausblatt 10}
\includegraphics[width=0.8\textwidth]{h10.pdf}
\subsection{Hausblatt 11}
\includegraphics[width=0.8\textwidth]{h11.pdf}
\subsection{Hausblatt 12}
\includegraphics[width=0.8\textwidth]{h12.pdf}
\subsection{Hausblatt 13}
\includegraphics[width=0.8\textwidth]{h13.pdf}
\subsection{Statistik über Übungsblätter}
\subsubsection{Boxplot}
\includegraphics[width=0.8\textwidth]{box.pdf}
\subsubsection{Violinplot}
\includegraphics[width=0.8\textwidth]{violin.pdf}

\subsection{Gesamtstatistik über alle Studierenden}

\subsubsection{Histogramm}
\includegraphics[width=0.8\textwidth]{totalpoints.pdf}

\subsubsection{Empirische Verteilungsfunktion}
\includegraphics[width=0.8\textwidth]{totalcumulative.pdf}


\end{document}
""")

    out = template.render(grades=formgrades,
                          succeeded=sorted(succeeded))
    return out 

############

def problems(grades, done_assigments):
    """Return  students who have achieved 
    less than needs_to_achieve assignments 
    at or above the needs_to_achieve 
    threshold. """

    needs_to_achieve = int(3/4*max_num_assignments)
    still_achieveable = max_num_assignments - done_assignments
    needs_to_have_achieved = max(needs_to_achieve - still_achieveable, 0)

    print("###", done_assignments, needs_to_achieve,
              still_achieveable, needs_to_have_achieved)
    
    failedpoints = 4
    failed = []
    endangered = []
    succeeded = [] # list of tuples (matrikel, bonus steps) 
    for matrikel, assignments in grades.items():
        # print(matrikel, assignments)
        a_over_threshold = sum([a['total'] >= failedpoints
                            for k, a
                            in assignments.items()])
        total_points = sum([a['total'] 
                            for k, a
                            in assignments.items()])
        # print(g['matrikelnr'], a_over_threshold)
        if a_over_threshold < needs_to_have_achieved:
            failed.append((matrikel,
                           a_over_threshold,
                           # assignments[1]['mail']
                           # we just need to grab any mail,
                           # hopefully, they are all the same: 
                           next(iter (assignments.values()))['mail'],
                           total_points
                               ))
        if a_over_threshold == needs_to_have_achieved:
            endangered.append((matrikel,
                           a_over_threshold,
                           # assignments[1]['mail']
                           # we just need to grab any mail,
                           # hopefully, they are all the same: 
                           next(iter (assignments.values()))['mail'],
                           total_points
                                   ))

        if a_over_threshold >= needs_to_achieve:
            succeeded.append((matrikel, total_points))
        
    return failed, endangered, succeeded 


    

############

def histogram(grades, done_assignments):
    data = [[], [], [], [],[], [],[], [],[], [],[], [],[], [],] 
    
    for matrikel, assignments in grades.items():
        for anr, a in assignments.items():
            # print(a['assignment'], a['total'])
            data[anr].append(a['total'])

    for i, d in enumerate(data[1:]):
        f = plt.figure()
        n, bins, patches = plt.hist(d, bins=21, range=(0,21))
        # print(n, bins)
        plt.xlabel('Punkte')
        plt.ylabel('Anzahl Blätter')
        plt.grid(True)
        plt.title('Hausblatt {}'.format(i+1))
        # plt.show()
        f.savefig('h{}.pdf'.format(i+1))

        

    # let's try a violin plot:
    # filter out zeros; they distort the plot
    no_zeros = list(map (lambda d: list(filter(lambda x: x> 0, d)), data)) 
    f = plt.figure()
    plt.violinplot(no_zeros[1:done_assignments+1])
    f.savefig('violin.pdf')

    # and a conventional boxplot
    f = plt.figure()
    plt.boxplot(no_zeros[1:done_assignments+1])
    f.savefig('box.pdf')

    
def total_histogram(f, e, s):

    total_points = ([x[3] for x in f] + 
                    [x[3] for x in e] +
                    [x[1] for x in s] )

    # pp(total_points)


    f = plt.figure()
    n, bins, patches = plt.hist(total_points, bins=20*13+1, range=(0,20*13+1))
    plt.xlabel('Punkte')
    plt.ylabel('Anzahl Studierende')
    plt.title('Verteilung der Gesamtpunkte auf Studierende -- Histogram')
    plt.grid(True)
    f.savefig('totalpoints.pdf')
    
    f = plt.figure()
    n, bins, patches = plt.hist(total_points, bins=20*13+1, range=(0,20*13+1),
                                    cumulative=True,
                                    histtype='step',
                                    normed=1)
    plt.xlabel('Punkte')
    plt.ylabel('Anteil Studierende mit dieser Gesamtpunktzahl')
    plt.title('Verteilung der Gesamtpunkte auf Studierende -- ECDF')
    plt.grid(True)
    f.savefig('totalcumulative.pdf')
    
    from scipy.stats import normaltest

    s, p = normaltest(total_points)
    print("normality:", s, p)


    # let's write out a total points statistic: (buggy, needs fixing)
    # with open('totalpoints.txt', 'w') as f:
    #     f.write('\n'.join(['{},{}'.format(x[0], x[3])
    #                            for x in f
    #                            ]))

    
############

def create_excel(grades):
    wb = Workbook()
    ws = wb.active

    # Create column heads:
    ws['A2'] = 'Matrikelnummer'
    ws['B2'] = 'Gesamtpunkte'
    ws['C2'] = '#H > Threshold'
    for i in range(13):
        ws.cell(column=4*i+4, row=1, value="HUE {}".format(i+1))
        ws.cell(column=4*i+4, row=2, value="Autograded")
        ws.cell(column=4*i+5, row=2, value="Manual")
        ws.cell(column=4*i+6, row=2, value="Total")
        ws.cell(column=4*i+7, row=2, value="email")

    for r, matrikel in enumerate(sorted(grades)):
        ws.cell(column=1, row=r+3, value=matrikel)

        ws.cell(column=2, row=r+3,
                    value=sum([int(a['total'])
                                   for k, a in grades[matrikel].items()]))
        ws.cell(column=3, row=r+3,
                    value=sum([int(int(a['total']) >= 4)
                                   for k, a in grades[matrikel].items()]))
        for c, akey in enumerate(grades[matrikel]):
            a = grades[matrikel][akey]
            ws.cell(column=4*c+4, row=r+3, value=a['point'])
            ws.cell(column=4*c+5, row=r+3, value=a['manual'])
            ws.cell(column=4*c+6, row=r+3, value=a['total'])
            ws.cell(column=4*c+7, row=r+3, value=a['mail'])
            

        
    wb.save('grades.xlsx')    
    

############

def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("manual", )
    args = parser.parse_args()
    return args
    

############

# Main: 

if __name__ == "__main__":

    try:
        done_assignments = int(sys.argv[3])
    except Exception:
        done_assignments = 0
    
    manualgrades = getManual(sys.argv[1])
    formgrades = getFormgrades(sys.argv[2])

    merge(formgrades, manualgrades)

    escape(formgrades, ['group', 'mail'])

    # sanity check:
    for matrikel, assignments in formgrades.items():
        for aa, blatt in assignments.items():
            if blatt['total'] > 20:
                print("PROBLEM: ", matrikel, aa, blatt)
                
    # check for problem students:
    
    failedstudents, endangeredstudents, succeeededstudents = problems(formgrades,
                            done_assignments)
    print("-------------------------------------------------------")
    print("Student without Exam admission:")
    pp(failedstudents)

    emaillister = lambda students: ', '.join([x[2] for x in students if len(x[2]) > 0])
    print(emaillister(failedstudents))
    print("-------------------------------------------------------")
    print("Student with endangered Exam admission:")
    pp(endangeredstudents)
    print(emaillister(endangeredstudents))

    print("-------------------------------------------------------")
    print("Admitted students with points") 
    pp(succeeededstudents)

    # this one is only true at the very end: 
    with open('admitted-matrikel.txt', 'w') as af:
        admitted = '\n'.join(
            sorted([str(x[0]) for x in succeeededstudents] ))
        
        af.write(admitted)
    
    with open('grades.tex', 'w') as gf:
        gf.write(output(formgrades, succeeededstudents))

    create_excel(formgrades)

    histogram(formgrades, done_assignments)

    total_histogram(failedstudents, endangeredstudents, succeeededstudents)

    # admitted = admitted_bonus(formgrades, done_assignments)

