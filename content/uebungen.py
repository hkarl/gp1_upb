import sys
import pprint
from openpyxl import load_workbook


pp = pprint.PrettyPrinter(indent=4)

numvl = int(sys.argv[1])

lines = []

for line in sys.stdin:
    line = line.rstrip()
    if 'UEBUNG' in line:
        entry["uebungen"].append(line)
        continue

    entry = {'line': line,
              'uebungen': []}

    lines.append(entry)

numlines = len(lines)
slidespervl = int(numlines/numvl)+1

lectures = []

for vl in range(numvl):
    lectures.append(lines[vl*slidespervl:(vl+1)*slidespervl])

# pp.pprint(lectures)


### Tie it to the Excel file's separation of Uebungen

wb = load_workbook("../meetings/Termine.xlsx")


def get_vorlesungen(wb, sheet, startzeile, endzeile, index1, index2, index3):
    
    pue_sheet = wb[sheet]

    pues = []
    for row in range(startzeile, endzeile):
        pues.append({'num': pue_sheet.cell(column=1, row=row).value,
                    'vl': [pue_sheet.cell(column=index1, row=row).value,
                            pue_sheet.cell(column=index2, row=row).value,
                            pue_sheet.cell(column=index3, row=row).value,
                               ]    
                        })

    return pues 

statistics = []

for kind, startzeile, endzeile, indexes in ( ('PUE', 3, 17, (13, 14, 15)),
                        ('HUE', 3, 16, (13, 14, 15))
                            ):

    print("======================")
    stats = {'kind': kind,
             'blaetter': []}
    
    vorlesungen = get_vorlesungen(wb, kind, startzeile, endzeile, *indexes)


    for p in vorlesungen:
        print("----------")
        print(kind, p['num'], end='')
        topics = []
        for l in p['vl']:
            if l == None:
                continue
        
            topics.extend(['\n'.join(u['uebungen']) for u in lectures[l-1] if u['uebungen'] ])

        # pp.pprint(topics)
        num_topics = len(topics)
        print(' #:', num_topics)
        
        print('\n'.join(topics))

        stats['blaetter'].append(num_topics)
        
    statistics.append(stats)

print("======================")
pp.pprint(statistics)
    
    
# [   {'blaetter': [0, 0, 5, 11, 14, 14, 21, 4, 5, 2, 4, 2, 6, 5], 'kind': 'PUE'},
#     {'blaetter':  [5, 15, 10, 21, 14, 7, 2, 4, 2, 2, 4, 5, 2], 'kind': 'HUE'}]
